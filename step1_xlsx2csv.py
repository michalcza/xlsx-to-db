import openpyxl
import os
import glob
import sys
import datetime
import csv

# Path to the folder containing xlsx files
folder_path = r'C:\Users\Michal\Documents\GitHub\xlsx-to-db'


# Values to delete
values_to_delete = ['DESCRIPTION', 'ND', 'OD', 'SI', 'MO', 'AI', 'EOF', 'INCLUDE IN REPORT (1=TRUE)', 'DAILY MAX VALUE', 'DAILY MIN VALUE', 'MAX VALUE TIME', 'MIN VALUE TIME','STATION', 'POINT', 'CATEGORY', 'HIGH OPERATIONAL (80%)', 'HIGH EMERGENCY (90%)', 'HIGH REASON', 'LOW OPERATIONAL', 'LOW EMERGENCY', 'LOW REASON'
]

# Value to delete after reading
delete_after_value = 2204

def delete_columns(ws, values_to_delete):
    """Delete columns that contain any value in values_to_delete."""
    for col in reversed(range(1, ws.max_column + 1)):
        for cell in ws[get_column_letter(col)]:
            if cell.value in values_to_delete:
                ws.delete_cols(col)
                break

# Output file name
output_file = 'output_{}.txt'.format(datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S'))

# Loop through all xlsx files in the folder
try:
    with open(output_file, 'w') as f:
        for file_path in glob.glob(os.path.join(folder_path, '*.xlsx')):
            # Load the workbook
            wb = openpyxl.load_workbook(file_path)

            # Loop through all worksheets in the workbook except for the 'Data' worksheet
            for sheet_name in wb.sheetnames.copy():
                if sheet_name != 'Data':
                    # Delete the worksheet
                    del wb[sheet_name]

            # Select the worksheet you want to modify
            ws = wb['Data']

            # Find the columns to delete
            columns_to_delete = []
            for column in range(1, ws.max_column + 1):
                for row in range(1, ws.max_row + 1):
                    if ws.cell(row=row, column=column).value in values_to_delete:
                        columns_to_delete.append(column)
                        break

            # Delete the columns
            for column in sorted(columns_to_delete, reverse=True):
                ws.delete_cols(column)
                
            # Find the row number of the first cell containing the value you want to delete after
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=4).value == delete_after_value:
            
            # Delete all rows after this row
                    ws.delete_rows(row + 1, ws.max_row)
         
            # save workbook as CSV
            output_filename = os.path.splitext(file_path)[0] + '.csv'

            with open(os.path.join(folder_path, output_filename), 'w', newline='', encoding='utf-8') as f_csv:
                writer = csv.writer(f_csv)
                for row in ws.iter_rows():
                    writer.writerow([cell.value for cell in row])

            # Write the output for the file to the output file
            f.write(f"{file_path}: {len(columns_to_delete)} columns deleted.\n")

            # Print the output for the file
            print(f"{file_path}: {len(columns_to_delete)} columns deleted.")

except Exception as e:
    print(f"An error occurred: {e}")
    sys.exit(1)

sys.exit(0)
